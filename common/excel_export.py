from __future__ import annotations

from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from common.utils import ensure_dir

__all__ = ["export_to_excel"]

_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")


def _autofit_columns(sheet: Worksheet, min_width: int = 10, max_width: int = 40) -> None:
    """Resize each column to roughly fit its widest cell's contents.

    Args:
        sheet: The worksheet whose columns should be resized.
        min_width: The smallest column width to allow.
        max_width: The largest column width to allow, to avoid a single
            long value stretching a column excessively.
    """
    for column_cells in sheet.columns:
        lengths = (len(str(cell.value)) for cell in column_cells if cell.value is not None)
        length = max(lengths, default=0)
        width = max(min_width, min(length + 2, max_width))
        column_letter = get_column_letter(column_cells[0].column)
        sheet.column_dimensions[column_letter].width = width


def _write_sheet(
    workbook: Workbook,
    sheet_name: str,
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
    is_first_sheet: bool,
) -> None:
    """Write one worksheet with a styled header row and data rows.

    Args:
        workbook: The workbook to add the sheet to.
        sheet_name: Title for the new worksheet (Excel truncates to 31
            characters; long names are shortened automatically).
        headers: Column header labels.
        rows: An iterable of row value sequences, one per record.
        is_first_sheet: If ``True``, reuse the workbook's default active
            sheet instead of creating a new one.
    """
    safe_name = sheet_name[:31]
    sheet = workbook.active if is_first_sheet else workbook.create_sheet()
    sheet.title = safe_name

    for col_idx, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGNMENT

    for row_idx, row_values in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_values, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)

    sheet.freeze_panes = "A2"
    _autofit_columns(sheet)


def export_to_excel(
    output_path: Path,
    sheets: Mapping[str, Sequence[Mapping[str, Any]]],
) -> Path:
   
    if not sheets:
        raise ValueError("'sheets' must contain at least one sheet.")

    ensure_dir(output_path.parent)
    workbook = Workbook()

    for sheet_index, (sheet_name, records) in enumerate(sheets.items()):
        if not records:
            raise ValueError(f"Sheet '{sheet_name}' has no records to export.")
        headers = list(records[0].keys())
        rows = [[record.get(header, "") for header in headers] for record in records]
        _write_sheet(
            workbook,
            sheet_name=sheet_name,
            headers=headers,
            rows=rows,
            is_first_sheet=(sheet_index == 0),
        )

    workbook.save(output_path)
    return output_path